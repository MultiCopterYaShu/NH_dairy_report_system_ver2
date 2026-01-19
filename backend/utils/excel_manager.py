from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import uuid

def export_work_items_to_excel(work_items):
    """作業項目をExcelにエクスポート"""
    wb = Workbook()
    ws = wb.active
    ws.title = "作業項目"
    
    # ヘッダー行
    headers = [
        'UUID', 'レベル1', 'レベル2', 'レベル3', 'レベル4', 'チェックリスト', '手段',
        '属性', '目標工数（分）', '社内リードタイム', '社内リードタイムUUID',
        '社外リードタイム', '社外リードタイムUUID', '担当種別'
    ]
    
    # ヘッダーのスタイル
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 階層パスを取得する関数
    def get_hierarchy_path(item, all_items):
        """項目の階層パス（レベル1から順に）を取得"""
        path = []
        current_item = item
        while current_item:
            path.insert(0, current_item.get('name', ''))
            parent_id = current_item.get('parent_id')
            if parent_id:
                current_item = next((i for i in all_items if i.get('id') == parent_id), None)
            else:
                break
        return path
    
    # 最下層の項目を判定する関数（子項目が存在しない = 最下層）
    def is_leaf_item(item, all_items):
        """項目が最下層かどうかを判定（子項目が存在しない = 最下層）"""
        item_id = item.get('id')
        # 子項目が存在するかチェック
        has_children = any(i.get('parent_id') == item_id for i in all_items)
        return not has_children
    
    # 最下層の項目のみをフィルタリング
    items_list = work_items.get('items', []) if isinstance(work_items, dict) else work_items
    leaf_items = [item for item in items_list if is_leaf_item(item, items_list)]
    
    # 階層順で並べ替え（深さ優先、同階層内は元の順序を維持）
    # まず、元のインデックスを取得するマップを作成
    item_index_map = {item.get('id'): idx for idx, item in enumerate(items_list)}
    
    def get_hierarchy_key(item, all_items, index_map):
        """階層パスと元のインデックスをキーとして返す（並べ替え用）"""
        path = get_hierarchy_path(item, all_items)
        # 階層パスをタプルとして返す
        key_parts = []
        for i in range(4):  # 最大4階層
            if i < len(path):
                # 階層パスの各レベルと名前を使用
                key_parts.append((i, path[i]))
            else:
                key_parts.append((i, ''))
        # 元のインデックスも追加（同階層内の順序を維持）
        original_index = index_map.get(item.get('id'), 999999)
        key_parts.append(original_index)
        return tuple(key_parts)
    
    # 階層パスと元のインデックスで並べ替え
    leaf_items_sorted = sorted(leaf_items, key=lambda item: get_hierarchy_key(item, items_list, item_index_map))
    
    # データ行
    row_idx = 2
    for item in leaf_items_sorted:
        hierarchy_path = get_hierarchy_path(item, items_list)
        
        ws.cell(row=row_idx, column=1, value=item.get('id', ''))  # UUID
        
        # レベル1-4に階層パスを設定
        for level in range(1, 5):
            if level <= len(hierarchy_path):
                ws.cell(row=row_idx, column=level + 1, value=hierarchy_path[level - 1])
            else:
                ws.cell(row=row_idx, column=level + 1, value='')
        
        # 残りの列（新しい順番）
        col_idx = 6
        ws.cell(row=row_idx, column=col_idx, value='\n'.join(item.get('checklist', [])))  # チェックリスト
        col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value='\n'.join(item.get('method', [])))  # 手段
        col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=item.get('attribute', ''))  # 属性
        col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=item.get('target_minutes', ''))  # 目標工数
        col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value='あり' if item.get('internal_leadtime') else 'なし')  # 社内リードタイム
        col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=','.join(item.get('internal_leadtime_items', [])))  # 社内リードタイムUUID
        col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value='あり' if item.get('external_leadtime') else 'なし')  # 社外リードタイム
        col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=','.join(item.get('external_leadtime_items', [])))  # 社外リードタイムUUID
        col_idx += 1
        ws.cell(row=row_idx, column=col_idx, value=','.join(item.get('担当種別', [])))  # 担当種別
        
        row_idx += 1
    
    # 列幅の調整
    ws.column_dimensions['A'].width = 36  # UUID
    ws.column_dimensions['B'].width = 25  # レベル1
    ws.column_dimensions['C'].width = 25  # レベル2
    ws.column_dimensions['D'].width = 25  # レベル3
    ws.column_dimensions['E'].width = 25  # レベル4
    ws.column_dimensions['F'].width = 30  # チェックリスト
    ws.column_dimensions['G'].width = 30  # 手段
    ws.column_dimensions['H'].width = 15  # 属性
    ws.column_dimensions['I'].width = 15  # 目標工数
    ws.column_dimensions['J'].width = 15  # 社内リードタイム
    ws.column_dimensions['K'].width = 40  # 社内リードタイムUUID
    ws.column_dimensions['L'].width = 15  # 社外リードタイム
    ws.column_dimensions['M'].width = 40  # 社外リードタイムUUID
    ws.column_dimensions['N'].width = 20  # 担当種別
    
    return wb

def export_project_to_excel_detail(project, work_types, work_items_by_type, all_reports):
    """プロジェクトをExcelにエクスポート（詳細版：日付、工数、リードタイム日数）"""
    from datetime import datetime
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトのシートを削除
    
    # プロジェクトに関連する工程を取得
    work_type_ids = project.get('work_type_ids', [])
    if not work_type_ids:
        return wb
    
    # プロジェクトIDを取得
    project_id = project.get('id')
    
    # 各工程ごとにシートを作成
    for work_type_id in work_type_ids:
        work_type = next((wt for wt in work_types if wt['id'] == work_type_id), None)
        if not work_type:
            continue
        
        work_items = work_items_by_type.get(work_type_id, [])
        if not work_items:
            continue
        
        # シートを作成
        ws = wb.create_sheet(title=work_type['name'][:31])  # Excelのシート名は31文字まで
        
        # ヘッダー行: 作業項目、作業の日付、工数、社内リードタイム日数、社外リードタイム日数
        ws.cell(row=1, column=1, value='作業項目')
        ws.cell(row=1, column=2, value='作業の日付')
        ws.cell(row=1, column=3, value='工数（分）')
        ws.cell(row=1, column=4, value='社内リードタイム日数')
        ws.cell(row=1, column=5, value='社外リードタイム日数')
        
        # ヘッダーのスタイル
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 階層パスを取得する関数
        def get_hierarchy_path(item, all_items):
            """項目の階層パス（レベル1から順に）を取得"""
            path = []
            current_item = item
            while current_item:
                path.insert(0, current_item.get('name', ''))
                parent_id = current_item.get('parent_id')
                if parent_id:
                    current_item = next((i for i in all_items if i.get('id') == parent_id), None)
                else:
                    break
            return path
        
        # 最下層の項目を判定する関数
        def is_leaf_item(item, all_items):
            """項目が最下層かどうかを判定"""
            item_id = item.get('id')
            has_children = any(i.get('parent_id') == item_id for i in all_items)
            return not has_children
        
        # 最下層の項目のみをフィルタリング（元の順序を維持）
        # work_itemsの順序を保持しながら、最下層項目のみを抽出
        item_index_map = {item.get('id'): idx for idx, item in enumerate(work_items)}
        leaf_items_with_index = []
        for item in work_items:
            if is_leaf_item(item, work_items):
                leaf_items_with_index.append((item_index_map[item.get('id')], item))
        
        # 元のインデックス順でソート（登録順を維持）
        leaf_items_sorted = [item for _, item in sorted(leaf_items_with_index, key=lambda x: x[0])]
        
        # 日報データを整理（作業項目ID -> レコードのリスト）
        # レコード: {date, username, minutes, work_item_id}
        report_records = []  # [{work_item_id, date, username, minutes}]
        
        for report in all_reports:
            username = report.get('username')
            if not username or username == 'admin':
                continue
            
            # このプロジェクトの日報か確認
            projects = report.get('projects', [])
            project_data = next((p for p in projects if p.get('project_id') == project_id), None)
            if not project_data:
                continue
            
            report_date = report.get('date')
            if not report_date:
                continue
            
            # このプロジェクトの作業項目を確認
            work_items_in_report = project_data.get('work_items', [])
            for work_item in work_items_in_report:
                work_item_id = work_item.get('work_item_id')
                work_item_type_id = work_item.get('work_type_id')
                
                # この工程の作業項目のみを対象
                if work_item_id and work_item_type_id == work_type_id:
                    minutes = work_item.get('minutes', 0)
                    report_records.append({
                        'work_item_id': work_item_id,
                        'date': report_date,
                        'username': username,
                        'minutes': minutes
                    })
        
        # リードタイム対象項目の日付を取得する関数
        def get_leadtime_target_date(target_work_item_id, current_date, all_records):
            """リードタイム対象項目の日付を取得（現在の日付より前の最新の日付）"""
            # 同じプロジェクト内のレコードのみを対象
            target_dates = [
                r['date'] for r in all_records
                if r['work_item_id'] == target_work_item_id and r['date'] < current_date
            ]
            if not target_dates:
                return None
            return max(target_dates)  # 最新の日付
        
        # 日付の差分を計算する関数
        def calculate_date_diff(date1_str, date2_str):
            """日付の差分を日数で返す（date1 - date2）"""
            try:
                date1 = datetime.strptime(date1_str, '%Y-%m-%d')
                date2 = datetime.strptime(date2_str, '%Y-%m-%d')
                diff = (date1 - date2).days
                return diff if diff >= 0 else None
            except:
                return None
        
        # データ行
        row_idx = 2
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        for item in leaf_items_sorted:
            work_item_id = item.get('id')
            hierarchy_path = get_hierarchy_path(item, work_items)
            work_item_path = ' > '.join(hierarchy_path)
            
            # この作業項目のレコードを取得
            item_records = [r for r in report_records if r['work_item_id'] == work_item_id]
            
            if not item_records:
                # レコードがない場合は作業項目名のみ表示
                ws.cell(row=row_idx, column=1, value=work_item_path)
                row_idx += 1
            else:
                # 各レコードを行として出力
                for record in sorted(item_records, key=lambda x: x['date']):
                    # 作業項目
                    ws.cell(row=row_idx, column=1, value=work_item_path)
                    
                    # 作業の日付
                    date_cell = ws.cell(row=row_idx, column=2, value=record['date'])
                    date_cell.fill = green_fill
                    
                    # 工数
                    ws.cell(row=row_idx, column=3, value=record['minutes'])
                    
                    # 社内リードタイム日数
                    internal_leadtime_items = item.get('internal_leadtime_items', [])
                    internal_leadtime_days = None
                    if internal_leadtime_items and len(internal_leadtime_items) > 0:
                        target_item_id = internal_leadtime_items[0]  # 最初の項目
                        target_date = get_leadtime_target_date(target_item_id, record['date'], report_records)
                        if target_date:
                            internal_leadtime_days = calculate_date_diff(record['date'], target_date)
                    
                    ws.cell(row=row_idx, column=4, value=internal_leadtime_days if internal_leadtime_days is not None else '')
                    
                    # 社外リードタイム日数
                    external_leadtime_items = item.get('external_leadtime_items', [])
                    external_leadtime_days = None
                    if external_leadtime_items and len(external_leadtime_items) > 0:
                        target_item_id = external_leadtime_items[0]  # 最初の項目
                        target_date = get_leadtime_target_date(target_item_id, record['date'], report_records)
                        if target_date:
                            external_leadtime_days = calculate_date_diff(record['date'], target_date)
                    
                    ws.cell(row=row_idx, column=5, value=external_leadtime_days if external_leadtime_days is not None else '')
                    
                    row_idx += 1
        
        # 列幅の調整
        ws.column_dimensions['A'].width = 50  # 作業項目
        ws.column_dimensions['B'].width = 15  # 作業の日付
        ws.column_dimensions['C'].width = 12  # 工数
        ws.column_dimensions['D'].width = 18  # 社内リードタイム日数
        ws.column_dimensions['E'].width = 18  # 社外リードタイム日数
    
    return wb

def export_project_to_excel(project, work_types, work_items_by_type, all_reports):
    """プロジェクトをExcelにエクスポート（工程別にシートを作成）"""
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトのシートを削除
    
    # プロジェクトに関連する工程を取得
    work_type_ids = project.get('work_type_ids', [])
    if not work_type_ids:
        return wb
    
    # プロジェクトIDを取得
    project_id = project.get('id')
    
    # このプロジェクトで作業しているユーザーのみを取得（admin以外）
    users = set()
    for report in all_reports:
        username = report.get('username')
        if not username or username == 'admin':
            continue
        
        # このプロジェクトの日報があるか確認
        projects = report.get('projects', [])
        project_data = next((p for p in projects if p.get('project_id') == project_id), None)
        if project_data:
            users.add(username)
    
    users = sorted(list(users))
    
    # 各工程ごとにシートを作成
    for work_type_id in work_type_ids:
        work_type = next((wt for wt in work_types if wt['id'] == work_type_id), None)
        if not work_type:
            continue
        
        work_items = work_items_by_type.get(work_type_id, [])
        if not work_items:
            continue
        
        # シートを作成
        ws = wb.create_sheet(title=work_type['name'][:31])  # Excelのシート名は31文字まで
        
        # ヘッダー行: 作業項目、ユーザー名
        ws.cell(row=1, column=1, value='作業項目')
        for col_idx, username in enumerate(users, 2):
            ws.cell(row=1, column=col_idx, value=username)
        
        # ヘッダーのスタイル
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(users) + 2):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 階層パスを取得する関数
        def get_hierarchy_path(item, all_items):
            """項目の階層パス（レベル1から順に）を取得"""
            path = []
            current_item = item
            while current_item:
                path.insert(0, current_item.get('name', ''))
                parent_id = current_item.get('parent_id')
                if parent_id:
                    current_item = next((i for i in all_items if i.get('id') == parent_id), None)
                else:
                    break
            return path
        
        # 最下層の項目を判定する関数
        def is_leaf_item(item, all_items):
            """項目が最下層かどうかを判定"""
            item_id = item.get('id')
            has_children = any(i.get('parent_id') == item_id for i in all_items)
            return not has_children
        
        # 最下層の項目のみをフィルタリング（元の順序を維持）
        # work_itemsの順序を保持しながら、最下層項目のみを抽出
        item_index_map = {item.get('id'): idx for idx, item in enumerate(work_items)}
        leaf_items_with_index = []
        for item in work_items:
            if is_leaf_item(item, work_items):
                leaf_items_with_index.append((item_index_map[item.get('id')], item))
        
        # 元のインデックス順でソート（登録順を維持）
        leaf_items_sorted = [item for _, item in sorted(leaf_items_with_index, key=lambda x: x[0])]
        
        # 日報データを整理（作業項目ID -> ユーザー -> 日付のマッピング）
        report_map = {}  # {work_item_id: {username: [dates]}}
        project_id = project.get('id')
        
        for report in all_reports:
            username = report.get('username')
            if not username or username == 'admin':
                continue
            
            # このプロジェクトの日報か確認
            projects = report.get('projects', [])
            project_data = next((p for p in projects if p.get('project_id') == project_id), None)
            if not project_data:
                continue
            
            report_date = report.get('date')
            if not report_date:
                continue
            
            # このプロジェクトの作業項目を確認
            work_items_in_report = project_data.get('work_items', [])
            for work_item in work_items_in_report:
                work_item_id = work_item.get('work_item_id')
                work_item_type_id = work_item.get('work_type_id')
                
                # この工程の作業項目のみを対象
                if work_item_id and work_item_type_id == work_type_id:
                    if work_item_id not in report_map:
                        report_map[work_item_id] = {}
                    if username not in report_map[work_item_id]:
                        report_map[work_item_id][username] = []
                    report_map[work_item_id][username].append(report_date)
        
        # データ行
        row_idx = 2
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        for item in leaf_items_sorted:
            work_item_id = item.get('id')
            hierarchy_path = get_hierarchy_path(item, work_items)
            work_item_path = ' > '.join(hierarchy_path)
            
            # 作業項目のパスを表示
            ws.cell(row=row_idx, column=1, value=work_item_path)
            
            # 各ユーザーの日報データをセルに記入
            for col_idx, username in enumerate(users, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # この作業項目とユーザーに対応する日報があるか確認
                if work_item_id in report_map and username in report_map[work_item_id]:
                    dates = sorted(report_map[work_item_id][username])
                    cell.value = ', '.join(dates)  # 複数の日付がある場合はカンマ区切り
                    cell.fill = green_fill  # 緑色に塗りつぶし
                else:
                    cell.value = ''
            
            row_idx += 1
        
        # 列幅の調整
        ws.column_dimensions['A'].width = 50  # 作業項目
        for col_idx, username in enumerate(users, 2):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 20  # ユーザー列
    
    return wb

def import_work_items_from_excel(file_path):
    """Excelから作業項目をインポート"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # 実際のデータがある行を取得（空行まで読み込む）
    max_row = ws.max_row
    
    # ヘッダー行をスキップして、全行を読み込む
    imported_items = []
    # 階層パスから親子関係を構築するためのマップ（階層パス -> 項目ID）
    hierarchy_map = {}
    
    def get_or_create_parent(parent_hierarchy, parent_level):
        """親項目を取得または作成"""
        if not parent_hierarchy:
            return None
        
        parent_key = tuple(parent_hierarchy)
        if parent_key in hierarchy_map:
            return hierarchy_map[parent_key]
        
        # 親項目が存在しない場合は作成
        parent_name = parent_hierarchy[-1]
        parent_uuid = str(uuid.uuid4())
        
        # さらに上位の親を取得
        grandparent_id = None
        if parent_level > 1:
            grandparent_hierarchy = parent_hierarchy[:-1]
            grandparent_id = get_or_create_parent(grandparent_hierarchy, parent_level - 1)
        
        parent_item = {
            'id': parent_uuid,
            'name': parent_name,
            'level': parent_level,
            'parent_id': grandparent_id,
            'attribute': None,
            'target_minutes': None,
            'checklist': [],
            'method': [],
            'internal_leadtime': False,
            'external_leadtime': False,
            'internal_leadtime_items': [],
            'external_leadtime_items': [],
            '担当種別': [],
            'is_leaf': False
        }
        
        imported_items.append(parent_item)
        hierarchy_map[parent_key] = parent_uuid
        
        return parent_uuid
    
    # 行を順番に処理（全行を読み込む）
    for row_num in range(2, max_row + 1):
        try:
            row_values = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
            if not row_values:
                continue
            row = list(row_values[0]) if row_values else []
        except:
            # 行が存在しない場合はスキップ
            continue
        
        # 行の長さを確認（少なくとも14列必要）
        if len(row) < 14:
            # 行が短い場合は空文字で埋める
            row = list(row) + [''] * (14 - len(row))
        
        # レベル1-4から階層パスを取得（最低2層は入っているものとして処理）
        level1 = ''
        level2 = ''
        level3 = ''
        level4 = ''
        
        if len(row) > 1 and row[1] is not None:
            level1 = str(row[1]).strip()
        if len(row) > 2 and row[2] is not None:
            level2 = str(row[2]).strip()
        if len(row) > 3 and row[3] is not None:
            level3 = str(row[3]).strip()
        if len(row) > 4 and row[4] is not None:
            level4 = str(row[4]).strip()
        
        # 最下層の項目名を取得（空文字を除外、最低2層は必須）
        hierarchy = []
        if level1:
            hierarchy.append(level1)
        if level2:
            hierarchy.append(level2)
        if level3:
            hierarchy.append(level3)
        if level4:
            hierarchy.append(level4)
        
        # 最低2層は必須
        if len(hierarchy) < 2:
            continue
        
        item_name = hierarchy[-1]  # 最後の階層が項目名
        level = len(hierarchy)
        
        # UUIDが空またはNoneの場合は新規生成
        item_uuid_raw = row[0] if len(row) > 0 else None
        if item_uuid_raw and str(item_uuid_raw).strip():
            item_uuid = str(item_uuid_raw).strip()
        else:
            item_uuid = str(uuid.uuid4())
        
        # 親項目のUUIDを取得（階層パスから、存在しない場合は作成）
        parent_id = None
        if level > 1:
            parent_hierarchy = hierarchy[:-1]  # 最後の要素を除いた階層パス
            parent_id = get_or_create_parent(parent_hierarchy, level - 1)
        
        # チェックリストと手段を配列に変換（列インデックス: 0=UUID, 1-4=レベル, 5=チェックリスト, 6=手段）
        checklist = []
        if len(row) > 5:
            checklist = [line.strip() for line in str(row[5] or '').split('\n') if line.strip()]
        method = []
        if len(row) > 6:
            method = [line.strip() for line in str(row[6] or '').split('\n') if line.strip()]
        
        # 属性と目標工数（列インデックス: 7=属性, 8=目標工数）
        attribute = None
        if len(row) > 7:
            attribute = row[7] if row[7] else None
        target_minutes = None
        if len(row) > 8:
            target_minutes = int(row[8]) if row[8] and str(row[8]).isdigit() else None
        
        # リードタイム対象項目を配列に変換（列インデックス: 9=社内リードタイム, 10=社内リードタイムUUID, 11=社外リードタイム, 12=社外リードタイムUUID）
        internal_leadtime = False
        if len(row) > 9:
            internal_leadtime = str(row[9] or '') == 'あり'
        internal_leadtime_items = []
        if len(row) > 10:
            internal_leadtime_items = [id.strip() for id in str(row[10] or '').split(',') if id.strip()]
        
        external_leadtime = False
        if len(row) > 11:
            external_leadtime = str(row[11] or '') == 'あり'
        external_leadtime_items = []
        if len(row) > 12:
            external_leadtime_items = [id.strip() for id in str(row[12] or '').split(',') if id.strip()]
        
        # リードタイムが「あり」でUUIDが空の場合、一つ前の行の作業項目を自動設定
        if internal_leadtime and not internal_leadtime_items:
            # 一つ前の行（imported_itemsの最後の要素）を取得
            if imported_items:
                previous_item = imported_items[-1]
                # 一つ前の項目が最下層項目であることを確認（子項目がない）
                has_children = any(i.get('parent_id') == previous_item.get('id') for i in imported_items)
                if not has_children:
                    internal_leadtime_items = [previous_item.get('id')]
        
        if external_leadtime and not external_leadtime_items:
            # 一つ前の行（imported_itemsの最後の要素）を取得
            if imported_items:
                previous_item = imported_items[-1]
                # 一つ前の項目が最下層項目であることを確認（子項目がない）
                has_children = any(i.get('parent_id') == previous_item.get('id') for i in imported_items)
                if not has_children:
                    external_leadtime_items = [previous_item.get('id')]
        
        # 担当種別を配列に変換（列インデックス: 13=担当種別）
        担当種別 = []
        if len(row) > 13:
            担当種別 = [cat.strip() for cat in str(row[13] or '').split(',') if cat.strip()]
        
        item = {
            'id': item_uuid,
            'name': item_name,
            'level': level,
            'parent_id': parent_id,
            'attribute': attribute,
            'target_minutes': target_minutes,
            'checklist': checklist,
            'method': method,
            'internal_leadtime': internal_leadtime,
            'external_leadtime': external_leadtime,
            'internal_leadtime_items': internal_leadtime_items,
            'external_leadtime_items': external_leadtime_items,
            '担当種別': 担当種別,
            'is_leaf': True  # インポートされる行は全て最下層
        }
        
        imported_items.append(item)
        
        # 階層マップに登録
        hierarchy_key = tuple(hierarchy)
        hierarchy_map[hierarchy_key] = item_uuid
    
    return imported_items

def export_project_view_to_excel(work_types, projects, work_items_by_type, all_reports):
    """プロジェクト別表示をExcelにエクスポート（工程ごとにシート、作業項目を行、プロジェクトを列）"""
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトのシートを削除
    
    # 工程ごとにシートを作成
    for work_type in work_types:
        work_type_id = work_type['id']
        work_items = work_items_by_type.get(work_type_id, [])
        if not work_items:
            continue
        
        # この工程に関連するプロジェクトを取得
        projects_with_work_type = [p for p in projects if p.get('work_type_ids') and work_type_id in p['work_type_ids']]
        if not projects_with_work_type:
            continue
        
        # シートを作成
        ws = wb.create_sheet(title=work_type['name'][:31])  # Excelのシート名は31文字まで
        
        # ヘッダー行: 作業項目、プロジェクト名
        ws.cell(row=1, column=1, value='作業項目')
        for col_idx, project in enumerate(projects_with_work_type, 2):
            ws.cell(row=1, column=col_idx, value=project['name'])
        
        # ヘッダーのスタイル
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(projects_with_work_type) + 2):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 階層パスを取得する関数
        def get_hierarchy_path(item, all_items):
            """項目の階層パス（レベル1から順に）を取得"""
            path = []
            current_item = item
            while current_item:
                path.insert(0, current_item.get('name', ''))
                parent_id = current_item.get('parent_id')
                if parent_id:
                    current_item = next((i for i in all_items if i.get('id') == parent_id), None)
                else:
                    break
            return path
        
        # 最下層の項目を判定する関数
        def is_leaf_item(item, all_items):
            """項目が最下層かどうかを判定"""
            item_id = item.get('id')
            has_children = any(i.get('parent_id') == item_id for i in all_items)
            return not has_children
        
        # 最下層の項目のみをフィルタリング（元の順序を維持）
        # work_itemsの順序を保持しながら、最下層項目のみを抽出
        item_index_map = {item.get('id'): idx for idx, item in enumerate(work_items)}
        leaf_items_with_index = []
        for item in work_items:
            if is_leaf_item(item, work_items):
                leaf_items_with_index.append((item_index_map[item.get('id')], item))
        
        # 元のインデックス順でソート（登録順を維持）
        leaf_items_sorted = [item for _, item in sorted(leaf_items_with_index, key=lambda x: x[0])]
        
        # プロジェクトと作業項目のマッピングを作成（{projectId: {workItemId: [dates]}}）
        project_work_item_map = {}
        
        for report in all_reports:
            if not report.get('projects'):
                continue
            
            report_date = report.get('date')
            if not report_date:
                continue
            
            for project_data in report['projects']:
                project_id = project_data.get('project_id')
                if project_id not in [p['id'] for p in projects_with_work_type]:
                    continue
                
                if project_id not in project_work_item_map:
                    project_work_item_map[project_id] = {}
                
                if not project_data.get('work_items'):
                    continue
                
                for work_item in project_data['work_items']:
                    work_item_id = work_item.get('work_item_id')
                    work_item_type_id = work_item.get('work_type_id')
                    
                    # この工程の作業項目のみを対象
                    if work_item_id and work_item_type_id == work_type_id:
                        if work_item_id not in project_work_item_map[project_id]:
                            project_work_item_map[project_id][work_item_id] = []
                        project_work_item_map[project_id][work_item_id].append(report_date)
        
        # データ行
        row_idx = 2
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        for item in leaf_items_sorted:
            work_item_id = item.get('id')
            hierarchy_path = get_hierarchy_path(item, work_items)
            work_item_path = ' > '.join(hierarchy_path)
            
            # 作業項目のパスを表示
            ws.cell(row=row_idx, column=1, value=work_item_path)
            
            # 各プロジェクトの日報データをセルに記入
            for col_idx, project in enumerate(projects_with_work_type, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                project_id = project['id']
                
                # この作業項目とプロジェクトに対応する日報があるか確認
                if project_id in project_work_item_map and work_item_id in project_work_item_map[project_id]:
                    dates = sorted(project_work_item_map[project_id][work_item_id])
                    cell.value = ', '.join(dates)  # 複数の日付がある場合はカンマ区切り
                    cell.fill = green_fill  # 緑色に塗りつぶし
                else:
                    cell.value = ''
            
            row_idx += 1
        
        # 列幅の調整
        ws.column_dimensions['A'].width = 50  # 作業項目
        for col_idx, project in enumerate(projects_with_work_type, 2):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = 25  # プロジェクト列
    
    return wb
